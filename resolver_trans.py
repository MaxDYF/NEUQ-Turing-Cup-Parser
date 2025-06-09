from openpyxl import load_workbook
import json
import gzip
import urllib.request

# 从excel中读取数据，生成队伍json文件
# 从pta上拿到队伍在pta中的id，作为board.xcpcio.com中的队伍id
# 这样子不用再做映射，省事

# 读入配置文件
cfg=""
with open('./config/cfg.json', 'r') as f:
    cfg = json.load(f)
problemId=cfg['problemId']
cookie=cfg['cookie']

dict_ans={}
dict_ans['PTA姓名']='xhc i love you'

url="https://pintia.cn/api/problem-sets/"+problemId+"/members?page=0&limit=500&filter=%7B%22userGroupId%22%3A%220%22%2C%22keyword%22%3A%22%22%7D"
# 直接从浏览器复制的header
headers={
    "Accept": "application/json;charset=UTF-8",
    "Accept-Encoding":"gzip, deflate, br, zstd",
    "Accept-Language":"zh-CN",
    "Cache-Control": "no-cache",
    "Content-Type": "application/json;charset=UTF-8",
    "Cookie": cookie,
    "Eagleeye-Pappname": "eksabfi2cn@94d5b8dc408ab8d",
    "Eagleeye-Sessionid": "3wlLtwL4j0X5772F1ymv1jt2bv0O",
    "Eagleeye-Traceid": "3f793bf1171646202705810238ab8d",
    "Pragma": "no-cache",
    "Priority": "u=1, i",
    "Referer": "https://pintia.cn/problem-sets/"+problemId+"/members",
    "Sec-Ch-Ua": '"Microsoft Edge";v="125", "Chromium";v="125", "Not.A/Brand";v="24"',
    "Sec-Ch-Ua-Mobile": "?0",
    "Sec-Ch-Ua-Platform": '"Windows"',
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-origin",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0",
    "X-Lollipop": "fd0f671d1f29b275fdfb4df1999ef66a",
    "X-Marshmallow": ""
}

# 发送请求
req=urllib.request.Request(url,headers=headers)
response=urllib.request.urlopen(req)

# 解析请求
content = gzip.decompress(response.read())
data = content.decode('utf-8')
json_data = json.loads(data)
list=json_data['members']

for i in range(0,len(list)):
    dict_ans[list[i]['studentUser']['studentNumber']]=list[i]['studentUser']['userId']



url1="https://pintia.cn/api/problem-sets/"+problemId+"/members?page=1&limit=500&filter=%7B%22userGroupId%22%3A%220%22%2C%22keyword%22%3A%22%22%7D"
req1=urllib.request.Request(url1,headers=headers)
response1=urllib.request.urlopen(req1)
# 解析请求
content1 = gzip.decompress(response1.read())
data1 = content1.decode('utf-8')
json_data1 = json.loads(data1)
list1=json_data1['members']

for i in range(0,len(list1)):
    dict_ans[list1[i]['studentUser']['studentNumber']]=list1[i]['studentUser']['userId']


url2="https://pintia.cn/api/problem-sets/"+problemId+"/members?page=2&limit=500&filter=%7B%22userGroupId%22%3A%220%22%2C%22keyword%22%3A%22%22%7D"
req2=urllib.request.Request(url2,headers=headers)
response2=urllib.request.urlopen(req2)
# 解析请求
content2 = gzip.decompress(response2.read())
data2 = content2.decode('utf-8')
json_data2 = json.loads(data2)
list2=json_data2['members']

for i in range(0,len(list2)):
    dict_ans[list2[i]['studentUser']['studentNumber']]=list2[i]['studentUser']['userId']


url3="https://pintia.cn/api/problem-sets/"+problemId+"/members?page=3&limit=500&filter=%7B%22userGroupId%22%3A%220%22%2C%22keyword%22%3A%22%22%7D"
req3=urllib.request.Request(url3,headers=headers)
response3=urllib.request.urlopen(req3)
# 解析请求
content3 = gzip.decompress(response3.read())
data3 = content3.decode('utf-8')
json_data3 = json.loads(data3)
list3=json_data3['members']

for i in range(0,len(list3)):
    if ('studentUser' in list3[i]):
        dict_ans[list3[i]['studentUser']['studentNumber']]=list3[i]['studentUser']['userId']


url4="https://pintia.cn/api/problem-sets/"+problemId+"/members?page=4&limit=500&filter=%7B%22userGroupId%22%3A%220%22%2C%22keyword%22%3A%22%22%7D"
req4=urllib.request.Request(url4,headers=headers)
response4=urllib.request.urlopen(req4)
# 解析请求
content4 = gzip.decompress(response4.read())
data4 = content4.decode('utf-8')
json_data4 = json.loads(data4)
list4=json_data4['members']

for i in range(0,len(list4)):
    dict_ans[list4[i]['studentUser']['studentNumber']]=list4[i]['studentUser']['userId']



print("finish")

# 加载Excel文件
wb = load_workbook('team.xlsx', data_only=True)

# 选择工作表
ws = wb['Sheet1']

ans={}

flag=False
# 遍历行和列来读取数据
for row in ws.iter_rows(values_only=True):
    # 排除表格中第一行的列名行
    if flag is not True:
        flag=True
        continue
    if row[0] == None:
        break
    ansItem={}
    ansItem['team_id']=dict_ans[str(row[14])]
    ansItem['name']=row[10]
    ansItem['organization']=row[2]
    if row[1][0]=='*':
        ansItem['official']=False
        ansItem['unofficial']=True
    else:
        ansItem['official']=True
        ansItem['unofficial']=False
    if row[9]=="N":
        ansItem['girl']=False
    else:
        ansItem['girl']=True
    ansItem['group']=[]
    ansItem['coach']=row[11]
    ansItem['members']=[]
    ansItem['members'].append(row[5])
    ansItem['members'].append(row[6])
    ansItem['members'].append(row[7])
    ans[ansItem['team_id']]=ansItem


# 指定要保存的文件名
filename = "./data/team.json"

# 打开文件并写入JSON数据
with open(filename, 'w', encoding='utf-8') as f:
    json.dump(ans, f, ensure_ascii=False, indent=4)

print(f"数据已成功保存到 {filename}")